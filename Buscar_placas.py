from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from config import user,cnpj,password
import requests
import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from lxml import html
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import os
import pyperclip
from selenium.webdriver.common.keys import Keys


# Obtém o diretório do script atual
script_directory = os.path.dirname(os.path.abspath(__file__))

def consultar_dados_veiculo(placa, driver):
    driver.get('https://buscaplacas.com.br/?ref=raf11')
    ## Adicione os valores coletados na próxima linha da planilha
    #sheet.append([placa_veiculo, chassi, ano_fab, ano_mod, cod_int, num_env, marca, modelo, versao, data_venda, valor_de, valor_venda, valor_plus, empresa, vendedor, Nome, Cpf, RG, telefone, celular, Cpfcnpj_cliente, RgIe_cliente, Nome_cliente, Placa_cliente, Descricao_cliente, Ano_cliente, Valor_cliente])
    workbook = openpyxl.Workbook()

    # Selecione a planilha ativa (geralmente criada por padrão)
    sheet = workbook.active

    # Adicione cabeçalhos às colunas
    sheet['A1'] = 'Placa do Veículo'
    sheet['B1'] = 'Ano Fabricante'
    sheet['C1'] = 'Modelo'   
    sheet['D1'] = 'Ano modelo' 
    sheet['E1'] = 'Cor do carro'
    sheet['F1'] = 'Marca '
  

# # Carregue a lista de placas a partir da planilha lista_placas
lista_placas_workbook = openpyxl.load_workbook('.\Lista placas.xlsx')
lista_placas_sheet = lista_placas_workbook.active

# Loop através das linhas da planilha lista_placas e colete as informações usando Selenium
for row in lista_placas_sheet.iter_rows(min_row=2, values_only=True):
    placa = row[0]
    print("Placa sendo pesquisada:", placa)

        
    lista_placas_workbook = openpyxl.load_workbook('.\Lista placas.xlsx')
    lista_placas_sheet = lista_placas_workbook.active

    # Carregue o workbook (arquivo) existente
    workbook = openpyxl.load_workbook('.\placas_pesquisadas.xlsx')

    # Adicione o cabeçalho na planilha somente se ela estiver vazia
    if not workbook.sheetnames:
        sheet = workbook[placa]
        cabecalho = ['Placa do Veículo', 'Ano Fabricante', 'Ano Fabricação', 'Ano Modelo', 'Modelo', 'Cor do carro','Marca']
        sheet.append(cabecalho)

    

    # Crie uma lista para armazenar as placas
    placas_lista = []

    # Loop através das linhas da planilha lista_placas a partir da segunda linha (A2)
    for row in lista_placas_sheet.iter_rows(min_row=2, values_only=True):
        placa = row[0]
        placas_lista.append(placa)
        
        
def main():
    options = Options()
    # Add any desired options to 'options' if needed.

    service = Service(executable_path='./chromedriver.exe')
    driver = webdriver.Chrome(service=service, options=options)


    driver.get('https://buscaplacas.com.br/?ref=raf11')
    driver.maximize_window()
    
     # Inicializar a planilha
    workbook = Workbook()
    planilha = workbook.active
    cabecalho = ['Placa do Veículo', 'Ano modelo', 'Ano Fabricação', 'Cor', 'Modelo','Marca']
    planilha.append(cabecalho)
    
    for placa in placas_lista:
        print("Pesquisando a placa:", placa)
        
                
        # Se houver uma instância anterior do driver, encerre-a
        if driver:
            driver.quit()

        # Inicialize um novo driver e execute as etapas de pesquisa
        driver = webdriver.Chrome(service=service, options=options)       
        driver.get('https://buscaplacas.com.br/?ref=raf11')
        driver.maximize_window()
          
        pyperclip.copy(placa)
        time.sleep(2)
        elemento_input = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="name-3b55"]')))
        elemento_input.clear()  # Limpa o campo, caso ele já tenha algum valor
        elemento_input.send_keys(Keys.CONTROL + "v")
        time.sleep(2)
        WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="sec-eee3"]/div/div[1]/form/div[2]/a'))).click()
        

        time.sleep(10)       
       
        
        

        # Localizar o elemento pelo XPath
        element = driver.find_element(By.XPATH, '/html/body/section/div/div[1]/table/tbody/tr[7]/td[2]')        
        ano= element.text        
        # Imprimir a informação da placa do veículo
        print("Ano Modelo:", ano)
        time.sleep(2)
        
        # Localizar o elemento pelo XPath
        element = driver.find_element(By.XPATH, '/html/body/section/div/div[1]/table/tbody/tr[8]/td[2]')        
        ano_modelo= element.text        
        # Imprimir a informação da placa do veículo
        print("Ano fabricaçao:", ano_modelo)
        time.sleep(2)
        
        # Localizar o elemento pelo XPath
        element = driver.find_element(By.XPATH, '/html/body/section/div/div[1]/table/tbody/tr[4]/td[2]')
        Modelo = element.text

        print("Modelo:", Modelo)        
        time.sleep(2)
        # Localizar o elemento pelo XPath
        element = driver.find_element(By.XPATH, '/html/body/section/div/div[1]/table/tbody/tr[10]/td[2]')
        Cor = element.text
        print("Cor:", Cor)
        
        time.sleep(2)
        # Localizar o elemento pelo XPath
        element = driver.find_element(By.XPATH, '/html/body/section/div/div[1]/table/tbody/tr[3]/td[2]')
        Marca = element.text
        print("marca:", Marca)

        
        # Adicione o cabeçalho na planilha (apenas uma vez)
        #cabecalho = ['Placa do Veículo', 'Ano fabricação', 'Ano Modelo', 'Cor', 'Modelo']
        planilha = workbook.active
        #planilha.append(cabecalho)       
        nova_linha = ([placa, ano_modelo, ano, Cor,Modelo,Marca ])
        planilha.append(nova_linha)

        # Salve o arquivo da planilha
        workbook.save('placas_pesquisadas.xlsx')
       
       

        # Salve a planilha com a nova aba e os dados
        workbook.save('placas_pesquisadas.xlsx')


    print("Planilha criada e valores enviados com sucesso!")
    driver.quit()

     
    
if __name__ == "__main__":
    main()
